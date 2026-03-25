import re
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from pathlib import Path

# Updated mapping dictionary for specific fund names.
mapping_dict = {
    "vanguard eurostoxx 50 ucits etf": "VX5E.L",
    "vanguard euro stoxx 50 ucits etf": "VX5E.L",  # Handles both variations.
    "vanguard ftse japan ucits etf": "VJPN.L",
    "vanguard ftse emerging markets ucits etf": "VFEM.L",
    "vanguard ftse wrld hi div yld ucits etf": "VHYL.L",
    "vfem.xlon.gb": "VFEM.L",
    "vx5e.xlon.gb": "VX5E.L",
    "vapx.xlon.gb": "VAPX.L",
    "vjpn.xlon.gb": "VJPN.L",
    "vusa.xlon.gb": "VUSA.L",
    "vger.xlon.gb": "VGER.L",
    "verx.xlon.g": "VERX.L",
    "global equity income fund - accumulation": "VAGEIGA",
    "global equity fund - accumulation": "VAGLEGA",
    "global small-cap index fund - accumulation": "VANGMSA",
    "emerging markets stock index fund - accumulation": "VANEMPA",
    "pacific ex-japan stock index fund - accumulation": "VAPEJPA",
    "japan stock index fund - accumulation": "VANJISA",
    "vhyl.xlon.gb": "VHYL.L",
    "vdxx.xlon.gb": "VDXX.L",
    "vnrt.xlon.gb": "VNRT.L",
    "vanguard ftse dev asiapac xjpn ucits etf": "VAPX.L",
    "vanguard glbl momentum factor ucits etf": "VMOM.L",
    "vanguard s&p 500 ucits etf": "VUSA.L",
    "lifestrategy 20% equity fund - gross accumulation": "VGLS20A",
    "lifestrategy 40% equity fund - accumulation": "VGLS40A",
    "lifestrategy 80% equity fund - accumulation": "VGLS80A",
    "lifestrategy 100% equity fund - accumulation": "VGL100A",
    "lifestrategy 60% equity fund - accumulation": "VGLS60A",
    "ftse 100 index unit trust accumulation": "VAFTIGA",
    "u.s. equity index fund - accumulation": "VUSEIDA.L",
    "v3am.xlon.gb": "V3AM.L",
    "vmid.xlon.gb": "VMID.L",
    "vanguard germany all cap ucits etf": "VGER.L",
}


def get_mapped_symbol(details):
    """
    Searches the details (in lowercase) for any key in the mapping dictionary.
    If found, returns the corresponding symbol; otherwise returns None.
    """
    details_lower = details.lower()
    for key, mapped_sym in mapping_dict.items():
        if key in details_lower:
            # List of symbols that should not have ".L" appended automatically.
            no_append_list = [
                "vageiga",
                "vaglega",
                "vangmsa",
                "vanempa",
                "vapejpa",
                "vanjisa",
                "vmom.l",
                "vaftiga",
                "vuseida.l",
                "v3am.l",
            ]
            if (
                not mapped_sym.endswith(".L")
                and mapped_sym.lower() not in no_append_list
            ):
                mapped_sym += ".L"
            return mapped_sym
    return None


def extract_activity_type(details):
    """
    Determines activity type based on keywords in the details string.

    Prioritizes checking for "sold" and "bought" before dividend keywords.
    Mapping rules (case-insensitive):
      - If trimmed details equal "deposit for investment purchases", returns DEPOSIT.
      - "Account Fee for the period" -> FEE
      - "Deposit for investment purposes" or "Deposit via direct credit" -> DEPOSIT
      - "Single Personal Pension Contribution" or "Pension Transfer In" -> DEPOSIT
      - "ETF dealing fee" -> FEE
      - If "sold" is found, returns SELL.
      - Else if "bought" is found, returns BUY.
      - "Interest" -> INTEREST
      - "Withdrawal" -> WITHDRAWAL
      - Then if "div" or "dividend" is found, returns DIVIDEND.
      - Otherwise, returns UNKNOWN.
    """
    d_strip = details.strip().lower()
    if d_strip == "deposit for investment purchases":
        return "DEPOSIT"
    d_lower = details.lower()
    if "account fee for the period" in d_lower:
        return "FEE"
    if "deposit for investment purposes" in d_lower:
        return "DEPOSIT"
    if "deposit via direct credit" in d_lower:
        return "DEPOSIT"
    if "single personal pension contribution" in d_lower:
        return "DEPOSIT"
    if "pension contribution tax relief" in d_lower:
        return "DEPOSIT"
    if "pension transfer in" in d_lower:
        return "DEPOSIT"
    if "etf dealing fee" in d_lower:
        return "FEE"
    if "sold" in d_lower:
        return "SELL"
    if "bought" in d_lower:
        return "BUY"
    if "interest" in d_lower:
        return "INTEREST"
    if "withdrawal" in d_lower:
        return "WITHDRAWAL"
    if re.search(r"\bdiv\b", details, re.IGNORECASE) or "dividend" in d_lower:
        return "DIVIDEND"
    return "UNKNOWN"


def extract_symbol(details, activity_type):
    """
    Determines the symbol from the details string with the following priority:
      1. Search for all text within parentheses.
         - Exclude any candidate containing "buy" or "sell" (case-insensitive).
         - If valid candidates exist, choose the first and ensure it ends with ".L".
      2. If no valid candidate is found, check the mapping dictionary.
      3. For DEPOSIT, WITHDRAWAL, or INTEREST trades, force the symbol to "$CASH-GBP".
      4. If nothing is found, return NULL (NaN).
    """

    # Priority 1: Check mapping dictionary.
    mapped = get_mapped_symbol(details)
    if mapped is not None:
        return mapped

    # Priority 2: Look for text within parentheses.
    candidates = re.findall(r"\(([^)]+)\)", details)
    valid_candidates = []
    for cand in candidates:
        if "buy" in cand.lower() or "sell" in cand.lower():
            continue
        valid_candidates.append(cand.strip())
    if valid_candidates:
        sym = valid_candidates[0]
        if not sym.endswith(".L"):
            sym += ".L"
        return sym

    # Priority 3: For DEPOSIT, WITHDRAWAL, or INTEREST, force cash symbol.
    if activity_type in ["DEPOSIT", "WITHDRAWAL", "INTEREST"]:
        return "$CASH-GBP"

    # Priority 4: Nothing found; return NULL.
    return np.nan


def extract_quantity(details, activity_type):
    """
    Extracts the quantity from the details string.

    - For FEE rows, returns None (to be handled separately).
    - Otherwise, attempts to extract a numeric value following "Bought" or "Sold".
    """
    if activity_type == "FEE":
        return None
    match = re.search(r"(Bought|Sold)\s+([\d,\.]+)", details, re.IGNORECASE)
    if match:
        qty_str = match.group(2).replace(",", "")
        try:
            return float(qty_str)
        except ValueError:
            return None
    return None


def calculate_unit_price(amount, quantity):
    """
    Computes the unit price as amount divided by quantity.
    Returns the absolute value (ensuring unitPrice is nonnegative).
    Returns None if quantity is missing or zero.
    """
    try:
        if quantity and quantity != 0:
            return abs(amount / quantity)
        else:
            return None
    except Exception:
        return None


def convert_excel():
    """
    Docstring for convert_excel
    """

    #Select file to process or do nothing
    file_path = filedialog.askopenfilename(
        title="Select excel input file", filetypes=[("Excel Files", "*.xlsx"),]
    )
    if not file_path:
        print("No file selected. Exiting.")
        return
    
    #if file is found, clean and start returning dataframes
    wb = load_workbook(file_path)

    # Delete the worksheet "Summary" if it exists
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "summary":
            wb.remove(wb[sheet_name])

    # Dictionary to hold our final DataFrames
    dataframes = {}

    # Process all remaining worksheets
    for ws in wb.worksheets:
        # Remove rows 1 to 4
        ws.delete_rows(1, 4)

        # Find first instance of "Balance" in column A
        balance_row = None
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value is not None and str(cell_value).strip() == "Balance":
                balance_row = row_idx
                break

        # Delete "Balance" row and everything below it
        if balance_row is not None:
            rows_to_delete = ws.max_row - balance_row + 1
            ws.delete_rows(balance_row, rows_to_delete)

         # Read the modified worksheet as a pandas DataFrame
        data = list(ws.values)
        
        if data:
            # Assuming the new first row (after deleting rows 1-4) is the header
            columns = data[0]
            df = pd.DataFrame(data[1:], columns=columns)

            # Store the DataFrame using the worksheet title as the dictionary key
            dataframes[ws.title] = df
        else:
            # Fallback if the worksheet is completely empty
            pass
            
    #initiate convertion of dataframes
    for sheet_name, df in dataframes.items():
        convert(df, sheet_name)



def convert(df,given_name):
    """
    process dataframe
    """

    # try:
    #     # Read CSV using an encoding that can handle non-UTF8 characters.
    #     df = pd.read_csv(file_path, encoding="latin1")
    # except Exception as e:
    #     print("Error reading the CSV file:", e)
    #     return

    # Convert the Date column from "DD/MM/YYYY" to "YYYY-MM-DD".
    try:
        df["Date"] = pd.to_datetime(df["Date"], dayfirst=True).dt.strftime("%Y-%m-%d")
    except Exception as e:
        print("Error processing the Date column:", e)

    # Ensure Details column is text.
    df["Details"] = df["Details"].astype(str)

    # Clean and convert Amount column to numeric.
    if "Amount" in df.columns:
        df["Amount"] = df["Amount"].replace(r"[\£\$,]", "", regex=True)
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
        # Store raw amount values.
        df["rawAmount"] = df["Amount"]

    # Determine ActivityType from Details.
    df["ActivityType"] = df["Details"].apply(extract_activity_type)

    # New condition: if Details contain "Payment by Faster" and raw Amount is negative, set ActivityType to WITHDRAWAL.
    mask_payment_by_faster = df["Details"].str.contains(
        "Payment by Faster", case=False, na=False
    ) & (df["rawAmount"] < 0)
    if mask_payment_by_faster.any():
        df.loc[mask_payment_by_faster, "ActivityType"] = "WITHDRAWAL"

    # Force all Amounts to be positive.
    df["Amount"] = df["Amount"].abs()

    # Extract Symbol using updated function.
    df["Symbol"] = df.apply(
        lambda row: extract_symbol(row["Details"], row["ActivityType"]), axis=1
    )

    # Extract Quantity.
    df["Quantity"] = df.apply(
        lambda row: extract_quantity(row["Details"], row["ActivityType"]), axis=1
    )

    # Force Quantity = 1 for INTEREST rows.
    df.loc[df["ActivityType"] == "INTEREST", "Quantity"] = 1

    # Calculate unitPrice (Amount/Quantity) where applicable.
    df["unitPrice"] = df.apply(
        lambda row: calculate_unit_price(row["Amount"], row["Quantity"]), axis=1
    )

    # Set constant currency.
    df["currency"] = "GBP"

    # Create fee column: for FEE rows, fee = absolute Amount.
    df["fee"] = df.apply(
        lambda row: abs(row["Amount"]) if row["ActivityType"] == "FEE" else np.nan,
        axis=1,
    )

    # Special Handling for "Account Fee for the period".
    mask_account_fee = df["Details"].str.contains(
        "Account Fee for the period", case=False, na=False
    )
    if mask_account_fee.any():
        df.loc[mask_account_fee, "Symbol"] = "$CASH-GBP"
        df.loc[mask_account_fee, "Quantity"] = 1
        df.loc[mask_account_fee, "unitPrice"] = df.loc[mask_account_fee, "Amount"]
        # Amount remains unchanged.

    # Handling for other FEE rows (excluding account fee rows).
    mask_other_fee = (df["ActivityType"] == "FEE") & (~mask_account_fee)
    if mask_other_fee.any():
        df.loc[mask_other_fee, "Quantity"] = 1
        df.loc[mask_other_fee, "unitPrice"] = 1.0
        # Clear Amount.
        df.loc[mask_other_fee, "Amount"] = np.nan

    # For BUY and SELL rows: round unitPrice to 10 decimal places and clear Amount.
    mask_buy_sell = df["ActivityType"].isin(["BUY", "SELL"])
    if mask_buy_sell.any():
        df.loc[mask_buy_sell, "unitPrice"] = df.loc[mask_buy_sell, "unitPrice"].apply(
            lambda x: round(abs(x), 10) if pd.notnull(x) else np.nan
        )
        df.loc[mask_buy_sell, "Amount"] = np.nan

    # For WITHDRAWAL rows: force Amount to be positive and set Symbol.
    mask_withdrawal = df["ActivityType"] == "WITHDRAWAL"
    if mask_withdrawal.any():
        df.loc[mask_withdrawal, "Symbol"] = "$CASH-GBP"
        df.loc[mask_withdrawal, "Amount"] = df.loc[mask_withdrawal, "Amount"].abs()

    # Arrange final columns.
    final_columns = [
        "Date",
        "Symbol",
        "Quantity",
        "ActivityType",
        "unitPrice",
        "currency",
        "fee",
        "Amount",
    ]
    final_df = df[final_columns]

    desktop = Path.home() / "Desktop"
    filename = desktop / f"{given_name}.csv"
   
    try:
        final_df.to_csv(filename, index=False, na_rep="")
        print("Output CSV file created successfully:", filename)

    except Exception as e:
        print("Error writing the output CSV file:", e)

def main():
    # Open file dialog to select the input CSV file.
    root = tk.Tk()

    root.title("Vanguard Transactions converter")
    root.minsize(200, 200)
    root.maxsize(500, 500)
    root.geometry("300x300+50+50")
    root.resizable(False, False)

    tk.Label(root, text="Version 0.1 -- 25 March 2025").pack()

    tk.Button(root, text="Convert EXCEL file", command= convert_excel).pack(padx=30,pady=30)

    root.mainloop()

if __name__ == "__main__":
    main()
